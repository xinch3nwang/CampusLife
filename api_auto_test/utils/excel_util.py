# -*- coding: utf-8 -*-
import openpyxl
from utils.log_util import Log

# 初始化日志记录器
mylogger = Log(__name__).getlogger()


class ExcelReader:
    def __init__(self, excelpath, sheetname):
        """
        初始化 ExcelReader 类的实例。

        :param excelpath: Excel 文件的路径
        :param sheetname: 要操作的工作表名称
        """
        # 保存 Excel 文件的路径
        self.filename = excelpath
        # 加载指定路径的 Excel 文件，要求文件后缀为 .xlsx
        self.wb = openpyxl.load_workbook(excelpath)
        # 获取指定名称的工作表
        self.sh = self.wb[sheetname]

    # 按行读取指定 sheet 表单中所有的内容
    def read_all_data_line_by_line(self):
        """
        一行一行地获取指定工作表中的所有数据，并将其转换为字典列表。

        :return: 包含所有行数据的字典列表，每个字典代表一行数据，键为表头
        """
        # 按行获取工作表中的数据并转换为列表
        row_datas = list(self.sh.rows)
        # 用于存储工作表的表头信息
        titles = []
        # 获取第一行的数据作为表头
        for title in row_datas[0]:
            titles.append(title.value)
        # 用于存储所有行的数据
        testdatas = []
        # 从第二行开始遍历所有行数据
        for case in row_datas[1:]:
            data = []
            for cell in case:
                try:
                    # 尝试将单元格的值添加到数据列表中
                    data.append(cell.value)
                except Exception as e:
                    # 记录读取 Excel 行数据失败的错误信息
                    mylogger.error("读取Excel行数据失败：{}".format(e), exc_info=True)
            # 将表头和当前行的数据组合成字典(类似JSON格式)
            case_data = dict(list(zip(titles, data)))
            # print(case_data)
            # 将当前行的数据字典添加到总数据列表中
            testdatas.append(case_data)
        return testdatas

    # 按行读取指定 sheet 表单中指定列中的内容
    def read_all_data_column_by_column(self, columns):
        """
        获取指定工作表中指定列的信息。

        :param columns: 要获取的列号列表或元组，例如 [1, 2, 3] 表示获取第 1、2、3 列的数据
        :return: 包含指定列数据的字典列表，每个字典代表一行数据，键为指定列的表头
        """
        # 获取工作表的最大行数
        maxline = self.sh.max_row
        # 用于存储指定列的表头信息
        titles = []
        # 用于存储所有行的指定列数据
        alldatas = []
        # 从第一行开始遍历所有行
        for linenum in range(1, maxline + 1):
            if linenum != 1:  # 如果不是第一行
                # 临时存储当前行指定列的数据
                onelinedata = []
                # 遍历要获取的列号
                for column in columns:
                    # 获取指定单元格的值
                    one_cell_value = self.sh.cell(linenum, column).value
                    try:
                        # 尝试将单元格的值转换为 Python 对象并添加到数据列表中
                        onelinedata.append(eval(one_cell_value))
                    except Exception as e:
                        # 如果转换失败，直接添加单元格的原始值
                        onelinedata.append(one_cell_value)
                # 将表头和当前行指定列的数据组合成字典
                onelinedata_dict = dict(list(zip(titles, onelinedata)))
                # 将当前行的数据字典添加到总数据列表中
                alldatas.append(onelinedata_dict)
            else:
                # 如果是第一行，获取指定列的表头信息
                for column in columns:
                    titles.append(self.sh.cell(linenum, column).value)

        return alldatas

    # 向指定单元格中写入数据
    def write_content_to_row_column(self, row, column, content):
        """
        向指定工作表的指定单元格中写入内容。

        :param row: 单元格所在的行号
        :param column: 单元格所在的列号
        :param content: 要写入的内容，类型为字符串
        """
        try:
            # 向指定单元格写入内容
            self.sh.cell(row=row, column=column, value=content)
            # 保存修改后的 Excel 文件
            self.wb.save("test.xlsx")
        except Exception as e:
            # 记录写入 Excel 数据失败的错误信息
            mylogger.error("写入excel数据失败：{}".format(e), exc_info=True)

